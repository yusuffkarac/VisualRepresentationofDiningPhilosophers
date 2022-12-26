import flet as ft
import time
from flet import IconButton,icons,colors
from threading import Thread, Lock
import random
import time
from openpyxl import load_workbook

class visuall:
    def __init__(self,delay, page,number_of_philosophers=5, meal_size=2):
        self.meals = [meal_size for _ in range(number_of_philosophers)]
        self.chopsticks = [Lock() for _ in range(number_of_philosophers)]
        self.number_of_philosophers = number_of_philosophers
        self.page = page
        self.delay = delay

        self.page.window_height= 900
        self.page.window_width = 900
        

        self.page.padding=ft.padding.only(top=20)
        self.page.auto_scroll = True
    def philosopher(self, i):
        phl_size = 100
        
        if self.number_of_philosophers>6:
            phl_size = 80
        
        phl_img=ft.Image(
                    src="phl.png",width=phl_size,height=phl_size,
                    fit=ft.ImageFit.SCALE_DOWN,

                    )
        chopstick_img=ft.Image(
                    src="chopstick.png",width=40,height=40,
                    )
        top_padding = -30
        left_padding = 700
        self.bg=ft.Container(ft.Image(
                     src="bg6.png",width=600,height=600,
         ),padding = ft.padding.only(top=20,left=left_padding-(left_padding*60/100)),
                )
        #top
        if i==0:
            phl_cont = ft.Container(
                padding = ft.padding.only(top=-4.3*top_padding,left=left_padding-(left_padding*20/100)),
                content=phl_img
                )
            chopstick_cont = ft.Container(
                padding = ft.padding.only(top=-7*top_padding,left=left_padding-(left_padding*17/100)),
                content=chopstick_img
                )

        #bottom
        elif i==self.number_of_philosophers-1:

            phl_cont = ft.Container(
                padding = ft.padding.only(top=-1.5*top_padding,left=left_padding-(left_padding*30/100)),
                content=phl_img
                )
            chopstick_cont = ft.Container(
                padding = ft.padding.only(top=-top_padding-15,left=left_padding-(left_padding*25/100)),
                content=chopstick_img
                )
            

        #rights
        elif i%2==0:

            #right-top
            if i%4==2:
                phl_cont = ft.Container(
                padding = ft.padding.only(top=top_padding-30,left=left_padding-(left_padding*10/100)),
                content=phl_img
                )
                chopstick_cont = ft.Container(
                padding = ft.padding.only(top=top_padding+20,left=left_padding-(left_padding*15/100)),
                content=chopstick_img
                )

            #right-bottom
            else:
                phl_cont = ft.Container(
                padding = ft.padding.only(top=-top_padding-10,left=left_padding-(left_padding*10/100)),
                content=phl_img
                )
                chopstick_cont = ft.Container(
                padding = ft.padding.only(top=-top_padding-20,left=left_padding-(left_padding*15/100)),
                content=chopstick_img
                )
        

            
            
        #lefts
        else:
            phl_img.src="phl_mirror.png"
            ##left-bottom
            if i%4==3:
                
                phl_cont = ft.Container(
                    padding = ft.padding.only(top=-top_padding,right=-left_padding-(left_padding*10/100)),
                    
                    content=phl_img
                    )
                chopstick_cont = ft.Container(
                    padding = ft.padding.only(top=-0.5*top_padding,right=-(left_padding+(left_padding*35/100))),
                    content=chopstick_img
                    )
            #left-top
            else:
                phl_cont = ft.Container(
                    padding = ft.padding.only(top=top_padding-10,right=-left_padding-(left_padding*10/100)),
                    
                    content=phl_img
                    )
                chopstick_cont = ft.Container(
                    padding = ft.padding.only(top=-.1*top_padding,right=-(left_padding+(left_padding*35/100))),
                    content=chopstick_img
                    ) 
        
        

        ## run 1 time for each philosopher
        if i==0:
            
            self.page.add(
                (
                ft.Container( 
                content=ft.Stack(
                [self.bg,phl_cont,chopstick_cont],

                    )
            )  ))
        
        else:
            self.page.add(
                (
                ft.Container( 
                content=ft.Stack(
                [phl_cont,chopstick_cont
                ],

                    ),
                padding=ft.padding.only(top=-400,left=50)
            )  ))
            
            
        

        print("philosopher",i,"started")
        self.page.title = "Dining Philosophers"
        self.page.update()
        self.page.bgimage = "bg.png"
        self.page.scroll = "always"
        j = (i+1) % self.number_of_philosophers
       
        def random_sleep():
            time.sleep(random.random()+self.delay)

        while self.meals[i] > 0:
            random_sleep()
            phl_img.src = "phl_think_mirror.png" if i%2==1 and i!=0 and i!=self.number_of_philosophers-1 else "phl_think.png"

            self.page.update()
            
            if not self.chopsticks[i].locked():
                self.chopsticks[i].acquire()
                chopstick_img.src = "empty.png"
                random_sleep()
                time.sleep(random.random())
                self.page.update()
                if not self.chopsticks[j].locked():
                    self.chopsticks[j].acquire()
                    phl_img.src = "phl_eat_mirror.png" if i%2==1 and i!=0 and i!=self.number_of_philosophers-1 else "phl_eat.png"
                    
                    chopstick_img.src="empty.png"


                    self.page.update()
                    
                    random_sleep()
                    self.meals[i] -= 1
                    self.chopsticks[j].release()
                    self.chopsticks[i].release()

                    phl_img.src = "phl_think_mirror.png" if i%2==1 and i!=0 and i!=self.number_of_philosophers-1 else "phl_think.png"
                    chopstick_img.src="chopstick.png"
                    self.page.update()

                    
                else:
                    
                    phl_img.src = "phl_think_mirror.png" if i%2==1 and i!=0 and i!=self.number_of_philosophers-1 else "phl_think.png"

                    self.page.update()
                    self.chopsticks[i].release()


                    chopstick_img.src="chopstick.png"
                    self.page.update()

        self.page.add(ft.Container(
            ft.Text(
            value=f"I am philosopher {i} and I am full now, sushi is my favorite meal",
        ),alignment=ft.alignment.center
        ))
        self.page.update()

           
            
            
            
###################################################################################################################################################
            
        
    
def main(page:ft.Page):
    
    page.window_height= 900
    page.window_width = 900
    
    drp_phl= ft.Dropdown(
            hint_text="Number of Philosophers",
            options=[
                ft.dropdown.Option(4),
                ft.dropdown.Option(5),
                ft.dropdown.Option(6),
                ft.dropdown.Option(7),
                ft.dropdown.Option(8),
                ft.dropdown.Option(9),
                ft.dropdown.Option(10),

            ],
            border_radius=30,
            filled=True,
            border_color=colors.TRANSPARENT,
            bgcolor=colors.RED,
            color=colors.WHITE,
            focused_bgcolor=colors.RED,
        )
    txt_meals = ft.TextField(
            hint_text = "Amount of Meals",
            
            border_radius = 30,
            filled = True,
            border_color = colors.TRANSPARENT,
            bgcolor = colors.RED,
            color = colors.WHITE,
            focused_bgcolor = colors.RED,
        )
    
    def start_dinner(t):
        page.window_full_screen = True
        n = int(drp_phl.value)
        m = int(txt_meals.value)
        if is_eating_slow.value:
            delay = 2
        else:
            delay = 0

        dining_philosophers = visuall(delay,page,n,m)
        
        philosophers = [Thread(target=dining_philosophers.philosopher, args=(i,)) for i in range(n)]
        toc = time.time()

        
        for philosopher in philosophers:
            philosopher.start()
        time.sleep(0.1)
        for philosopher in philosophers:
            philosopher.join()
            print("philosopher joined")
        tic = time.time()
        dinner_time = tic-toc

        ## Excel sheet part
        wrkbk = load_workbook("dining_philosophers_statistics.xlsx")
        
        # to identify the active sheet
        sh = wrkbk.active

        # identify the number of occupied rows
        max_row = sh.max_row

        # to insert the data in the next row
        sh.cell(row=max_row+1, column=1).value = n
        sh.cell(row=max_row+1, column=2).value = m
        sh.cell(row=max_row+1, column=3).value = dinner_time
        sh.cell(row=2, column=4).value = f"=AVERAGE(A2:A{max_row+1})"
        sh.cell(row=2, column=5).value = f"=AVERAGE(B2:B{max_row+1})"
        sh.cell(row=2, column=6).value = f"=AVERAGE(C2:C{max_row+1})"


        # to save the file
        wrkbk.save("dining_philosophers_statistics.xlsx")


        
        print(f"Simulation took {dinner_time} seconds")
    is_eating_slow = ft.Checkbox(label="Eat slow", value=False)
    
    mealstart_btn = IconButton(icons.SET_MEAL,
            
            icon_color= colors.BLUE,
            icon_size= 45,
            on_click=start_dinner
            )
    
    header = ft.Container(ft.Text(value="DINING PHILOSOPHERS",size=50,weight="bold",color=colors.CYAN),alignment=ft.alignment.center)

    page.add(header)
    page.add(
        ft.Container(
        
        ft.Row([
            drp_phl,txt_meals,is_eating_slow,mealstart_btn
        ])
        ,
        padding=ft.padding.only(left=page.width/4)
        )
    )
                

        
    
ft.app(target=main,view=ft.WEB_BROWSER,assets_dir="assets")
